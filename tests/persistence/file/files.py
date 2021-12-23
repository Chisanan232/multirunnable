from multirunnable.persistence.file.files import File, CSVFormatter, XLSXFormatter, JSONFormatter, Checking

from .._data import Test_Data_List, Test_Error_Data_List, Test_JSON_Data

from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook, load_workbook
import pytest
import json
import csv
import os


Test_CSV_File_Path: str = str(Path("./for_testing.csv"))
Test_XLSX_File_Path: str = str(Path("./for_testing.xlsx"))
Test_JSON_File_Path: str = str(Path("./for_testing.json"))

Test_Mode: str = "a+"

Run_Open_Process_Flag: bool = False
Run_Write_Process_Flag: bool = False
Run_Close_Process_Flag: bool = False
Run_Procedure_List: List[str] = []
Run_Result_Data_List = []


@pytest.fixture(scope="function")
def csv_formatter() -> CSVFormatter:
    return CSVFormatter()


@pytest.fixture(scope="function")
def xlsx_formatter() -> XLSXFormatter:
    return XLSXFormatter()


@pytest.fixture(scope="function")
def json_formatter() -> JSONFormatter:
    return JSONFormatter()


class TestFiles:

    def test_checking(self):
        try:
            Checking()
        except RuntimeError as re:
            assert "All methods in class 'CheckingUtils' is static method, you shouldn't new this class." in str(re), f""
        else:
            assert False, f"It should raise a RuntimeError if you try to instantiate it."


    def test_save_as_csv(self, csv_formatter: CSVFormatter):
        _test_file_path = Test_CSV_File_Path
        TestFiles._save_process(_file_format=csv_formatter, _file_path=_test_file_path, _data=Test_Data_List)

        _exist_file = os.path.exists(_test_file_path)
        assert _exist_file is True, "It should exist a .csv file."
        with open(_test_file_path, "r") as csv_file:
            data = csv.reader(csv_file)
            for index, d in enumerate(data):
                for ele_d, ele_o in zip(d, Test_Data_List[index]):
                    assert str(ele_d) == str(ele_o), "Each values in the data row should be the same."

        os.remove(_test_file_path)


    def test_save_as_csv_with_invalid_data_format(self, csv_formatter: CSVFormatter):
        _test_file_path = Test_CSV_File_Path
        try:
            TestFiles._save_process(_file_format=csv_formatter, _file_path=_test_file_path, _data=Test_Error_Data_List)
        except ValueError as ve:
            assert True, f"It work finely."
        else:
            assert False, f"It should raise ValueError if the data which we save as target file format is invalid."
        os.remove(_test_file_path)


    def test_csv_format_stream(self, csv_formatter: CSVFormatter):
        _data_stream = csv_formatter.stream(data=Test_Data_List)
        _data_stream_list = _data_stream.split("\r\n")
        _data_stream_iter = filter(lambda a: a is not None and a != "", _data_stream_list)
        _filter_data_stream_list = list(_data_stream_iter)
        for i, _ds in enumerate(_filter_data_stream_list):
            _test_data_row = list(map(lambda a: str(a), Test_Data_List[i]))
            assert _ds == ",".join(_test_data_row)


    def test_save_as_xlsx(self, xlsx_formatter: XLSXFormatter):
        _test_file_path = Test_XLSX_File_Path
        TestFiles._save_process(_file_format=xlsx_formatter, _file_path=_test_file_path, _data=Test_Data_List)

        _exist_file = os.path.exists(_test_file_path)
        assert _exist_file is True, "It should exist a .xlsx file."

        _workbook = load_workbook(filename=_test_file_path, read_only=True)
        for _sheet in _workbook:
            for _row, _data_row in zip(_sheet.iter_rows(min_row=_sheet.min_row, max_row=_sheet.max_row, min_col=_sheet.min_column, max_col=_sheet.max_column), Test_Data_List):
                for _cell, _data_content in zip(_row, _data_row):
                    _value = _cell.value
                    assert _value == _data_content, f"Each values in the data row should be the same."

        os.remove(_test_file_path)


    def test_save_as_json(self, json_formatter: JSONFormatter):
        _test_file_path = Test_JSON_File_Path
        TestFiles._save_process(_file_format=json_formatter, _file_path=_test_file_path, _data=Test_JSON_Data)

        _exist_file = os.path.exists(_test_file_path)
        assert _exist_file is True, "It should exist a .json file."
        with open(_test_file_path, "r") as json_file:
            _data = json.load(json_file)
            assert "data" in _data.keys(), f"The key 'data' should be in the JSON format data."
            assert _data["data"] is not None and len(_data["data"]) != 0, f"It should has some data row with the key 'data' in JSON format content."
            for index, d in enumerate(_data["data"]):
                for ele_d, ele_o in zip(d, Test_Data_List[index]):
                    assert str(ele_d) == str(ele_o), "Each values in the data row should be the same."

        os.remove(_test_file_path)


    def test_json_format_stream(self, json_formatter: JSONFormatter):
        _data_stream = json_formatter.stream(data=Test_Data_List)
        for _d, _td in zip(json.loads(_data_stream), Test_Data_List):
            assert _d == _td, f"Each values in the data row should be the same."


    @staticmethod
    def _save_process(_file_format: File, _file_path, _data):
        try:
            _file_format.file_path = _file_path
            _file_format.mode = "a+"
            _file_format.encoding = "UTF-8"
            _file_format.open()
            _file_format.write(data=_data)
            _file_format.close()
        except ValueError as e:
            raise e
        except Exception as e:
            assert False, f"It should work finely without any issue."
        else:
            assert True, f"It work finely!"

        assert _file_format.file_path == _file_path, f"It should be '{_file_format.file_path}' we set."
        assert _file_format.mode == "a+", f"It should be '{_file_format.mode}' we set."
        assert _file_format.encoding == "UTF-8", f"It should be '{_file_format.encoding}' we set."

